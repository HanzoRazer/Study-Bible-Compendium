"""
Excel and CSV import helpers for the Study Bible Compendium.

This module:
- Opens .xlsx files via openpyxl or .csv files via csv module.
- Detects header row and column mapping.
- Yields normalized verse rows: (book, chapter, verse, text).
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterator, Optional, Tuple, List

from .util import info, warn


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


@dataclass
class ExcelVerseRow:
    book: str          # Book name or code
    chapter: int
    verse: int
    text: str
    raw_row_index: int  # for diagnostics


HEADER_CANDIDATES: Dict[str, List[str]] = {
    "book": ["book", "bookname", "book_name", "bk"],
    "chapter": ["chapter", "chap", "ch"],
    "verse": ["verse", "verse_num", "vs", "v"],
    "text": ["text", "verse_text", "content", "body"],
}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def _detect_column_mapping(headers: List[object]) -> Optional[Dict[str, int]]:
    """
    Try to find which column index corresponds to book/chapter/verse/text.
    Returns a mapping { 'book': idx, 'chapter': idx, 'verse': idx, 'text': idx }
    or None if detection fails.
    """
    norm_headers = [_normalize_header(h) for h in headers]
    mapping: Dict[str, int] = {}

    for logical_name, candidates in HEADER_CANDIDATES.items():
        idx_found: Optional[int] = None
        for i, norm in enumerate(norm_headers):
            if any(norm == cand for cand in candidates):
                idx_found = i
                break
        if idx_found is None:
            warn(f"Could not detect column for '{logical_name}'. Headers were: {headers}")
            return None
        mapping[logical_name] = idx_found

    return mapping


def iter_verses_from_excel(
    excel_path: Path,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
) -> Iterator[ExcelVerseRow]:
    """
    Yield ExcelVerseRow objects from Excel (.xlsx) or CSV (.csv) files.

    Parameters
    ----------
    excel_path:
        Path to the Excel or CSV file.
    sheet_name:
        Optional worksheet name (Excel only). If None, the active sheet is used.
    max_rows:
        Optional limit on number of data rows yielded (for testing).

    Yields
    ------
    ExcelVerseRow instances.
    """
    excel_path = excel_path.resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")
    
    # Check file extension to determine handler
    suffix = excel_path.suffix.lower()
    
    if suffix == '.csv':
        yield from _iter_verses_from_csv(excel_path, max_rows)
    elif suffix in ('.xlsx', '.xlsm', '.xls'):
        yield from _iter_verses_from_xlsx(excel_path, sheet_name, max_rows)
    else:
        warn(f"Unsupported file format: {suffix}. Expected .csv, .xlsx, .xlsm, or .xls")
        return


def _iter_verses_from_csv(
    csv_path: Path,
    max_rows: Optional[int] = None,
) -> Iterator[ExcelVerseRow]:
    """Handle CSV file import."""
    info(f"Opening CSV file: {csv_path}")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        
        try:
            headers = next(reader)
        except StopIteration:
            warn("CSV file is empty.")
            return
        
        info(f"Detected header row: {headers}")
        mapping = _detect_column_mapping(headers)
        if mapping is None:
            warn("Failed to detect required columns; aborting CSV import.")
            return
        
        count = 0
        for row_idx, row in enumerate(reader, start=2):  # 1-based, +1 for header
            if max_rows is not None and count >= max_rows:
                info(f"Stopping after max_rows={max_rows} rows.")
                break
            
            if len(row) < max(mapping.values()) + 1:
                warn(f"Row {row_idx}: not enough columns; skipping.")
                continue
            
            try:
                book_raw = row[mapping["book"]]
                chapter_raw = row[mapping["chapter"]]
                verse_raw = row[mapping["verse"]]
                text_raw = row[mapping["text"]]
            except IndexError:
                warn(f"Row {row_idx}: column access error; skipping.")
                continue
            
            if not book_raw or not chapter_raw or not verse_raw:
                warn(f"Row {row_idx}: missing book/chapter/verse; skipping.")
                continue
            
            text_str = text_raw.strip() if text_raw else ""
            if not text_str:
                warn(f"Row {row_idx}: empty verse text; skipping.")
                continue
            
            try:
                chapter_int = int(chapter_raw)
                verse_int = int(verse_raw)
            except (TypeError, ValueError):
                warn(f"Row {row_idx}: non-integer chapter/verse; skipping. "
                     f"chapter={chapter_raw!r}, verse={verse_raw!r}")
                continue
            
            book_str = book_raw.strip()
            if not book_str:
                warn(f"Row {row_idx}: empty book value; skipping.")
                continue
            
            yield ExcelVerseRow(
                book=book_str,
                chapter=chapter_int,
                verse=verse_int,
                text=text_str,
                raw_row_index=row_idx,
            )
            count += 1


def _iter_verses_from_xlsx(
    excel_path: Path,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
) -> Iterator[ExcelVerseRow]:
    """Handle Excel file import."""
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Install it with: pip install openpyxl"
        )

    excel_path = excel_path.resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    info(f"Opening Excel file: {excel_path}")
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)

    if sheet_name is None:
        ws = wb.active
        info(f"Using active sheet: {ws.title!r}")
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        info(f"Using sheet: {ws.title!r}")

    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        warn("Excel sheet is empty.")
        return

    headers_list = list(headers)
    info(f"Detected header row: {headers_list}")
    mapping = _detect_column_mapping(headers_list)
    if mapping is None:
        warn("Failed to detect required columns; aborting Excel import.")
        return

    count = 0
    for row_idx, row in enumerate(rows, start=2):  # 1-based row index; +1 for header
        if max_rows is not None and count >= max_rows:
            info(f"Stopping after max_rows={max_rows} rows.")
            break

        row_list = list(row)

        try:
            book_raw = row_list[mapping["book"]]
            chapter_raw = row_list[mapping["chapter"]]
            verse_raw = row_list[mapping["verse"]]
            text_raw = row_list[mapping["text"]]
        except IndexError:
            warn(f"Row {row_idx}: not enough columns; skipping.")
            continue

        if book_raw is None or chapter_raw is None or verse_raw is None:
            warn(f"Row {row_idx}: missing book/chapter/verse; skipping.")
            continue

        text_str = "" if text_raw is None else str(text_raw).strip()
        if not text_str:
            # Allow empty text? For now, skip.
            warn(f"Row {row_idx}: empty verse text; skipping.")
            continue

        try:
            chapter_int = int(chapter_raw)
            verse_int = int(verse_raw)
        except (TypeError, ValueError):
            warn(f"Row {row_idx}: non-integer chapter/verse; skipping. "
                 f"chapter={chapter_raw!r}, verse={verse_raw!r}")
            continue

        book_str = str(book_raw).strip()
        if not book_str:
            warn(f"Row {row_idx}: empty book value; skipping.")
            continue

        yield ExcelVerseRow(
            book=book_str,
            chapter=chapter_int,
            verse=verse_int,
            text=text_str,
            raw_row_index=row_idx,
        )
        count += 1
