#!/usr/bin/env python3
"""
Convert Excel files (Page, Line, Text) from PDF extraction into normalized CSV (book, chapter, verse, text).

This tool parses verses from the Text column using regex patterns like:
  "Genesis 1:1 In the beginning..."
  "John 3:16 For God so loved..."

Usage:
    python tools/convert_excel_to_csv.py --input sources/excel/excel_output/kjv.xlsx --output data/converted/kjv.csv
    python tools/convert_excel_to_csv.py --batch sources/excel/excel_output/ --outdir data/converted/
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("[error] openpyxl not found. Install: pip install openpyxl")
    sys.exit(1)


# Book/chapter header pattern (e.g., "Genesis 1", "John 3")
HEADER_PATTERN = re.compile(r"^([A-Za-z0-9\s]+?)\s+(\d+)$")

# Verse split pattern (matches digit followed by text at word boundary)
VERSE_SPLIT_PATTERN = re.compile(r"(\d+)(?=[A-Z])")

# Common Bible book names
BIBLE_BOOKS = {
    "Genesis", "Exodus", "Leviticus", "Numbers", "Deuteronomy",
    "Joshua", "Judges", "Ruth", "1 Samuel", "2 Samuel", "1 Kings", "2 Kings",
    "1 Chronicles", "2 Chronicles", "Ezra", "Nehemiah", "Esther",
    "Job", "Psalm", "Psalms", "Proverbs", "Ecclesiastes", "Song of Solomon",
    "Isaiah", "Jeremiah", "Lamentations", "Ezekiel", "Daniel",
    "Hosea", "Joel", "Amos", "Obadiah", "Jonah", "Micah", "Nahum",
    "Habakkuk", "Zephaniah", "Haggai", "Zechariah", "Malachi",
    "Matthew", "Mark", "Luke", "John",
    "Acts", "Romans", "1 Corinthians", "2 Corinthians", "Galatians",
    "Ephesians", "Philippians", "Colossians",
    "1 Thessalonians", "2 Thessalonians", "1 Timothy", "2 Timothy",
    "Titus", "Philemon", "Hebrews", "James",
    "1 Peter", "2 Peter", "1 John", "2 John", "3 John",
    "Jude", "Revelation"
}


def normalize_book_name(raw: str) -> Optional[str]:
    """Normalize book name variations to canonical form."""
    raw = raw.strip()
    # Try exact match first
    if raw in BIBLE_BOOKS:
        return raw
    # Try case-insensitive
    for book in BIBLE_BOOKS:
        if book.lower() == raw.lower():
            return book
    # Handle abbreviations (extend as needed)
    abbrev_map = {
        "Gen": "Genesis", "Exod": "Exodus", "Lev": "Leviticus",
        "Num": "Numbers", "Deut": "Deuteronomy", "Josh": "Joshua",
        "Judg": "Judges", "1Sam": "1 Samuel", "2Sam": "2 Samuel",
        "1Kgs": "1 Kings", "2Kgs": "2 Kings", "Ps": "Psalms",
        "Prov": "Proverbs", "Eccl": "Ecclesiastes", "Song": "Song of Solomon",
        "Isa": "Isaiah", "Jer": "Jeremiah", "Lam": "Lamentations",
        "Ezek": "Ezekiel", "Dan": "Daniel", "Matt": "Matthew",
        "Rom": "Romans", "1Cor": "1 Corinthians", "2Cor": "2 Corinthians",
        "Gal": "Galatians", "Eph": "Ephesians", "Phil": "Philippians",
        "Col": "Colossians", "1Thess": "1 Thessalonians", "2Thess": "2 Thessalonians",
        "1Tim": "1 Timothy", "2Tim": "2 Timothy", "Heb": "Hebrews",
        "Jas": "James", "1Pet": "1 Peter", "2Pet": "2 Peter",
        "1Jn": "1 John", "2Jn": "2 John", "3Jn": "3 John",
        "Rev": "Revelation"
    }
    return abbrev_map.get(raw)


def parse_book_chapter_header(text: str) -> Optional[Tuple[str, int]]:
    """
    Extract (book, chapter) from header line like "Genesis 1" or "1 Samuel 12".
    Returns None if line doesn't match pattern.
    """
    match = HEADER_PATTERN.match(text.strip())
    if not match:
        return None
    
    raw_book = match.group(1).strip()
    chapter_str = match.group(2)
    
    book = normalize_book_name(raw_book)
    if not book:
        return None
    
    try:
        chapter = int(chapter_str)
    except ValueError:
        return None
    
    return (book, chapter)


def split_verses(text: str) -> list[Tuple[int, str]]:
    """
    Split text with inline verse numbers like:
      "1In the beginning God created...2And the earth was..."
    
    Returns list of (verse_number, verse_text) tuples.
    """
    verses = []
    parts = VERSE_SPLIT_PATTERN.split(text)
    
    # parts will be like: ['', '1', 'In the beginning...', '2', 'And the earth...']
    i = 1
    while i < len(parts):
        if i + 1 < len(parts):
            try:
                verse_num = int(parts[i])
                verse_text = parts[i + 1].strip()
                if verse_text:
                    verses.append((verse_num, verse_text))
            except ValueError:
                pass
        i += 2
    
    return verses


def parse_verse_line(text: str) -> Optional[Tuple[str, int, int, str]]:
    """
    DEPRECATED: Old single-line verse parser.
    Kept for compatibility but not used in new implementation.
    """
    return None



def convert_excel_to_csv(
    excel_path: Path,
    csv_path: Path,
    verbose: bool = False
) -> None:
    """
    Convert single Excel file from (Page, Line, Text) to (book, chapter, verse, text).
    
    This parser is stateful:
    1. Tracks current book/chapter from headers like "Genesis 1"
    2. Splits text by inline verse numbers (e.g., "1In...2And...")
    3. Accumulates verse text across multiple rows
    """
    print(f"[info] Converting {excel_path.name}...")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # Find Text column
    header = [cell.value for cell in ws[1]]
    try:
        text_col_idx = header.index("Text")
    except ValueError:
        print(f"[warn] No 'Text' column in {excel_path.name}, skipping")
        wb.close()
        return
    
    verses = []
    current_book = None
    current_chapter = None
    pending_verse_text = []  # Accumulate text for current verse
    pending_verse_num = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= text_col_idx:
            continue
        
        text = row[text_col_idx]
        if not text or not isinstance(text, str):
            continue
        
        text = text.strip()
        
        # Check if this is a book/chapter header
        header_match = parse_book_chapter_header(text)
        if header_match:
            # Save any pending verse
            if current_book and current_chapter and pending_verse_num:
                verses.append((
                    current_book,
                    current_chapter,
                    pending_verse_num,
                    " ".join(pending_verse_text)
                ))
                pending_verse_text = []
                pending_verse_num = None
            
            current_book, current_chapter = header_match
            if verbose:
                print(f"  ...found {current_book} {current_chapter}")
            continue
        
        # Skip if we haven't found a book/chapter yet
        if not current_book or not current_chapter:
            continue
        
        # Try to extract verses from this line
        extracted_verses = split_verses(text)
        
        if extracted_verses:
            # Save any pending verse from previous line
            if pending_verse_num:
                verses.append((
                    current_book,
                    current_chapter,
                    pending_verse_num,
                    " ".join(pending_verse_text)
                ))
                pending_verse_text = []
            
            # Add new verses (except possibly the last one which might continue)
            for i, (v_num, v_text) in enumerate(extracted_verses):
                if i < len(extracted_verses) - 1:
                    # Complete verse
                    verses.append((current_book, current_chapter, v_num, v_text))
                else:
                    # Last verse might continue on next line
                    pending_verse_num = v_num
                    pending_verse_text = [v_text]
            
            if verbose and len(verses) % 1000 == 0:
                print(f"  ...parsed {len(verses)} verses")
        else:
            # No verse markers, might be continuation of previous verse
            if pending_verse_num:
                pending_verse_text.append(text)
    
    # Save final pending verse
    if current_book and current_chapter and pending_verse_num:
        verses.append((
            current_book,
            current_chapter,
            pending_verse_num,
            " ".join(pending_verse_text)
        ))
    
    wb.close()
    
    # Write CSV
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["book", "chapter", "verse", "text"])
        writer.writerows(verses)
    
    print(f"[ok] Converted {len(verses)} verses to {csv_path.name}")



def batch_convert(
    input_dir: Path,
    output_dir: Path,
    verbose: bool = False
) -> None:
    """
    Convert all .xlsx files in input_dir to CSV in output_dir.
    """
    excel_files = list(input_dir.glob("*.xlsx"))
    
    if not excel_files:
        print(f"[warn] No .xlsx files found in {input_dir}")
        return
    
    print(f"[info] Found {len(excel_files)} Excel files")
    
    for excel_path in sorted(excel_files):
        csv_path = output_dir / excel_path.with_suffix(".csv").name
        try:
            convert_excel_to_csv(excel_path, csv_path, verbose)
        except Exception as e:
            print(f"[error] Failed to convert {excel_path.name}: {e}")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Convert Excel (Page, Line, Text) to normalized CSV (book, chapter, verse, text)"
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Single Excel file to convert"
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output CSV path (for single file mode)"
    )
    parser.add_argument(
        "--batch",
        type=Path,
        help="Batch convert all .xlsx files in directory"
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        help="Output directory for batch mode"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Verbose output"
    )
    
    args = parser.parse_args(argv)
    
    if args.input and args.output:
        # Single file mode
        convert_excel_to_csv(args.input, args.output, args.verbose)
    elif args.batch and args.outdir:
        # Batch mode
        batch_convert(args.batch, args.outdir, args.verbose)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
